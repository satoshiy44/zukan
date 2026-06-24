#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
台本スプレッドシート生成スクリプト

入力: videos.json （1本=1エントリ。画像パスと台本フィールドを持つ）
出力:
  - 台本スプレッドシート.xlsx … 画像を埋め込み、台本フィールドを横に並べたシート（推敲用）
  - 台本一覧.md             … GitHub上でも確認できるMarkdown版

使い方:
  python3 build_spreadsheet.py [videos.json] [出力ディレクトリ]
  （省略時: ./videos.json を読み、このスクリプトの2つ上=「コマ切り抜き動画/」直下に出力）
"""
import json
import os
import sys

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from PIL import Image as PILImage

THUMB_W = 200  # 埋め込みサムネ幅(px)

# (見出し, jsonキー, 列幅(文字))
COLUMNS = [
    ("管理No", "no", 8),
    ("上コマ(起)", None, 30),
    ("下コマ(転)", None, 30),
    ("作品名/作者", "title", 22),
    ("質問文", "question", 22),
    ("上コマ セリフ", "top_text", 22),
    ("下コマ セリフ", "bottom_text", 22),
    ("オチ/引き", "hook", 22),
    ("背景", "bg", 10),
    ("出す秒数", "reveal_sec", 9),
    ("表示秒数", "hold_sec", 9),
    ("狙い", "goal", 12),
    ("FANZAリンク", "link", 26),
    ("推敲メモ", "note", 26),
    ("OK", "ok", 6),
]

HEADER_FILL = PatternFill("solid", fgColor="2F5597")
HEADER_FONT = Font(color="FFFFFF", bold=True)
WRAP = Alignment(wrap_text=True, vertical="center", horizontal="left")
CENTER = Alignment(vertical="center", horizontal="center")


def thumb(path, base_dir):
    """画像を読み、アスペクト比維持でTHUMB_W幅にしたサイズ(px)を返す。"""
    p = path if os.path.isabs(path) else os.path.join(base_dir, path)
    with PILImage.open(p) as im:
        w, h = im.size
    scale = THUMB_W / float(w)
    return p, THUMB_W, int(h * scale)


def build(videos, base_dir, out_dir):
    wb = Workbook()
    ws = wb.active
    ws.title = "台本"
    ws.freeze_panes = "A2"

    # ヘッダー
    for ci, (head, _key, width) in enumerate(COLUMNS, start=1):
        c = ws.cell(row=1, column=ci, value=head)
        c.fill, c.font, c.alignment = HEADER_FILL, HEADER_FONT, CENTER
        ws.column_dimensions[get_column_letter(ci)].width = width
    ws.row_dimensions[1].height = 22

    md = ["# 台本一覧", "", "推敲用。OK列が埋まったものから動画化に進めます。", ""]

    for ri, v in enumerate(videos, start=2):
        max_img_h = 80
        for ci, (head, key, _w) in enumerate(COLUMNS, start=1):
            col = get_column_letter(ci)
            if head in ("上コマ(起)", "下コマ(転)"):
                imgkey = "top_img" if head == "上コマ(起)" else "bottom_img"
                src = v.get(imgkey)
                if src:
                    try:
                        p, w, h = thumb(src, base_dir)
                        xi = XLImage(p)
                        xi.width, xi.height = w, h
                        ws.add_image(xi, f"{col}{ri}")
                        max_img_h = max(max_img_h, h)
                    except Exception as e:
                        ws.cell(row=ri, column=ci, value=f"[画像読込失敗] {src}\n{e}").alignment = WRAP
            else:
                val = v.get(key, "")
                cell = ws.cell(row=ri, column=ci, value=val)
                cell.alignment = CENTER if ci == 1 else WRAP
        ws.row_dimensions[ri].height = max_img_h * 0.78  # px→pt 目安

        # Markdown版
        md.append(f"## {v.get('no','')}  {v.get('title','')}")
        if v.get("top_img"):
            md.append(f"- 上コマ: `{v['top_img']}`")
        if v.get("bottom_img"):
            md.append(f"- 下コマ: `{v['bottom_img']}`")
        md += [
            f"- 質問文: {v.get('question','')}",
            f"- 上セリフ: {v.get('top_text','')}",
            f"- 下セリフ: {v.get('bottom_text','')}",
            f"- オチ/引き: {v.get('hook','')}",
            f"- 背景: {v.get('bg','')} / 出す: {v.get('reveal_sec','')}s / 表示: {v.get('hold_sec','')}s",
            f"- 狙い: {v.get('goal','')} / リンク: {v.get('link','')}",
            "",
        ]

    os.makedirs(out_dir, exist_ok=True)
    xlsx_path = os.path.join(out_dir, "台本スプレッドシート.xlsx")
    md_path = os.path.join(out_dir, "台本一覧.md")
    wb.save(xlsx_path)
    with open(md_path, "w", encoding="utf-8") as f:
        f.write("\n".join(md))
    print("出力:", xlsx_path)
    print("出力:", md_path)


def main():
    json_path = sys.argv[1] if len(sys.argv) > 1 else os.path.join(os.path.dirname(__file__), "videos.json")
    here = os.path.dirname(os.path.abspath(__file__))
    out_dir = sys.argv[2] if len(sys.argv) > 2 else os.path.dirname(here)  # = コマ切り抜き動画/
    with open(json_path, encoding="utf-8") as f:
        videos = json.load(f)
    base_dir = os.path.dirname(os.path.abspath(json_path))
    build(videos, base_dir, out_dir)


if __name__ == "__main__":
    main()
